import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

def create_spreadsheet(file_path):
    if not os.path.exists(file_path):
        headers = ['Date', 'Description', 'Transaction', 'Debit', 'Credit', 'Balance', 'Reconciled']
        df = pd.DataFrame(columns=headers)
        df.to_excel(file_path, index=False)

def add_transaction(file_path, date, description, transaction, debit, credit, reconciled=False):
    try:
        df = pd.read_excel(file_path)
    except FileNotFoundError:
        df = pd.DataFrame(columns=['Date', 'Description', 'Transaction', 'Debit', 'Credit', 'Balance', 'Reconciled'])

    balance = df['Balance'].iloc[-1] + credit - debit if len(df) > 0 else credit - debit
    new_txn = pd.DataFrame([{
        'Date': date,
        'Description': description,
        'Transaction': transaction,
        'Debit': debit,
        'Credit': credit,
        'Balance': balance,
        'Reconciled': reconciled
    }])
    df = pd.concat([df, new_txn], ignore_index=True)
    df.to_excel(file_path, index=False)
    apply_formatting(file_path)

def carry_balance_to_next_month(file_path):
    df = pd.read_excel(file_path)
    last_balance = df['Balance'].iloc[-1]
    last_date = pd.to_datetime(df['Date'].iloc[-1])
    next_month_date = (last_date + pd.offsets.MonthBegin(1)).strftime('%Y-%m-%d')
    new_txn = pd.DataFrame([{
        'Date': next_month_date,
        'Description': 'Starting Balance',
        'Transaction': '',
        'Debit': 0.0,
        'Credit': 0.0,
        'Balance': last_balance,
        'Reconciled': False
    }])
    df = pd.concat([df, new_txn], ignore_index=True)
    df.to_excel(file_path, index=False)

def apply_formatting(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    red_font = Font(color="FF0000")
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value < 0:
                cell.font = red_font
    wb.save(file_path)

def generate_monthly_report(file_path, month, year):
    df = pd.read_excel(file_path)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce', infer_datetime_format=True)
    monthly_df = df[(df['Date'].dt.month == month) & (df['Date'].dt.year == year)]
    total_debit = monthly_df['Debit'].sum()
    total_credit = monthly_df['Credit'].sum()
    ending_balance = monthly_df['Balance'].iloc[-1] if not monthly_df.empty else 0.0
    pdf_file_path = f"data/monthly_report_{year}_{month}.pdf"
    c = canvas.Canvas(pdf_file_path, pagesize=letter)
    c.drawString(100, 750, f"Monthly Report for {month}/{year}")
    c.drawString(100, 730, f"Total Debit: {total_debit}")
    c.drawString(100, 710, f"Total Credit: {total_credit}")
    c.drawString(100, 690, f"Ending Balance: {ending_balance}")
    c.save()

def generate_yearly_report(file_path, year):
    df = pd.read_excel(file_path)
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce', infer_datetime_format=True)
    yearly_df = df[df['Date'].dt.year == year]
    total_debit = yearly_df['Debit'].sum()
    total_credit = yearly_df['Credit'].sum()
    ending_balance = yearly_df['Balance'].iloc[-1] if not yearly_df.empty else 0.0
    pdf_file_path = f"data/yearly_report_{year}.pdf"
    c = canvas.Canvas(pdf_file_path, pagesize=letter)
    c.drawString(100, 750, f"Yearly Report for {year}")
    c.drawString(100, 730, f"Total Debit: {total_debit}")
    c.drawString(100, 710, f"Total Credit: {total_credit}")
    c.drawString(100, 690, f"Ending Balance: {ending_balance}")
    c.save()

def export_transactions_to_csv(file_path):
    df = pd.read_excel(file_path)
    csv_file_path = file_path.replace('.xlsx', '.csv')
    df.to_csv(csv_file_path, index=False)

def reconcile_transaction(file_path, transaction_id):
    df = pd.read_excel(file_path)
    df.loc[df['Transaction'] == transaction_id, 'Reconciled'] = True
    df.to_excel(file_path, index=False)